#! python
# Send emails based on oayment status in spreadsheet
import openpyxl, smtplib, sys

# open the spreadsheet and getget the latest dues status.
wb = openpyxl.load_workbook('duesRecord.xlsx')
sheet = wb.get_sheet_by_name('sheet1')
lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row=1, column=lastCol).value
# Check each member's payment status
unpaidMembers = {}
for i in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMemmbers[name] = email
        
# Log in to email account
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('my_email_address@gmail.com', 'my_pass_key')

# Send out reminder emails.
for name, email in unpaidMembers.items():
    body = ("Subject: %s dues unpaid. \nDear %s, \nRecords show that you haven't paid dues for %s. please make this payment as soon as possible. Thank you!" %(latestMonth, name, latestMonth))
    print("Sending email to %s..." % email)
    sendmailStatus = smtpObj.sendmail("my_email_address@gmail.com", email, body)
    
    if sendemailStatus != {}:
        print("There was a problem sending email to %s: %s" % (email, sendemailStatus))
smtpObj.quit()

